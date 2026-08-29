"""
File parsing for bulk data import (CSV / Excel).

Deliberately minimal: reads an uploaded file fully into memory (bounded
by MAX_FILE_SIZE_BYTES), returns a plain list of row dicts keyed by
header name, plus the detected header list. No pandas — the built-in
`csv` module and `openpyxl` are enough for this scope and keep the
dependency footprint small.
"""

import csv
import io

from fastapi import UploadFile
from openpyxl import load_workbook

from app.utils.exceptions import InvalidFileError

ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024  # 50 MB, per the Module 7 spec


def _get_extension(filename: str) -> str:
    if "." not in filename:
        return ""
    return "." + filename.rsplit(".", 1)[-1].lower()


def _stringify_cell(value) -> str:
    """Normalize any cell value (str, number, None, datetime) to a trimmed string."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    # utf-8-sig handles a leading BOM, which Excel adds when it exports CSVs.
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = [h.strip() for h in (reader.fieldnames or [])]
    rows = []
    for raw_row in reader:
        rows.append({key.strip(): _stringify_cell(value) for key, value in raw_row.items() if key})
    return headers, rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    row_iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(row_iterator)
    except StopIteration:
        return [], []

    headers = [_stringify_cell(h) for h in header_row]
    rows = []
    for raw_row in row_iterator:
        # Skip fully blank trailing rows, which openpyxl sometimes yields.
        if raw_row is None or all(cell is None for cell in raw_row):
            continue
        row = {}
        for header, value in zip(headers, raw_row):
            if header:
                row[header] = _stringify_cell(value)
        rows.append(row)
    return headers, rows


async def parse_uploaded_file(file: UploadFile) -> tuple[list[str], list[dict[str, str]]]:
    """
    Validate and parse an uploaded CSV/Excel file.

    Returns (headers, rows). Raises InvalidFileError for a disallowed
    extension or a file over MAX_FILE_SIZE_BYTES — callers turn that
    into a clean 400/413, never a raw parsing exception.
    """
    extension = _get_extension(file.filename or "")
    if extension not in ALLOWED_EXTENSIONS:
        raise InvalidFileError(
            f"Unsupported file type '{extension or 'unknown'}'. Only .csv and .xlsx files are allowed."
        )

    content = await file.read()
    if len(content) > MAX_FILE_SIZE_BYTES:
        raise InvalidFileError(
            f"File is too large ({len(content) / (1024 * 1024):.1f} MB). Maximum allowed size is 50 MB."
        )
    if len(content) == 0:
        raise InvalidFileError("The uploaded file is empty.")

    if extension == ".csv":
        headers, rows = _parse_csv(content)
    else:
        try:
            headers, rows = _parse_xlsx(content)
        except Exception as exc:  # openpyxl raises various error types for corrupt files
            raise InvalidFileError(
                "Could not read this Excel file — it may be corrupted or not a valid .xlsx file."
            ) from exc

    if not headers:
        raise InvalidFileError("No header row was found in the uploaded file.")

    return headers, rows
